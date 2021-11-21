import openpyxl
from openpyxl.styles import Font
import datetime
import pytz
from tzlocal import get_localzone
from gui import *


# FUNCTION CURRENTLY NOT IN USE
def create_empty_file(file: str, task: str, time: str) -> None:
    """
    Creates empty .xlsx file with predefined column names and sets
    the font to Arial Narrow size 10,5pt.

    :param file: file name
    :param task: name for column B, where task names will be stored
    :param time: name for column C, where task times will be stored
    """
    global arial_narrow_font
    global arial_narrow_bold_font

    workbook = openpyxl.Workbook()
    empty_sheet = workbook.active

    # set name, font and width for column B,
    # where task names will be stored
    empty_sheet.column_dimensions['B'].width = 35
    empty_sheet.column_dimensions['B'].font = arial_narrow_font
    empty_sheet['B2'] = task
    empty_sheet['B2'].font = arial_narrow_bold_font

    # set name, font and width for column C,
    # where task times will be stored
    empty_sheet.column_dimensions['C'].width = 10
    empty_sheet.column_dimensions['C'].font = arial_narrow_font
    empty_sheet['C2'] = time
    empty_sheet['C2'].font = arial_narrow_bold_font

    workbook.save(file)
    workbook.close()


def find_cell(ws: openpyxl.worksheet, column_range: int, row_range: int,
              name: str) -> openpyxl.cell:
    """Find cell in Excel sheet with `name` value.

    :param ws: sheet to be searched through
    :param column_range: maximum non-empty column
    :param row_range: maximum non-empty row
    :param name: value to search for
    :return: cell with value `name`. `None` if it's not found.
    """
    for column in range(1, column_range + 1):
        for row in range(1, row_range + 1):
            temp_cell = ws.cell(row=row, column=column)
            if temp_cell.value == name:
                return temp_cell
    return None


def write_data(file: str, task: str, time: int, task_label: str,
               time_label: str) -> None:
    """
    Add task and its time (in hours) to Excel sheet. Return from the
    function if file name is not specified (empty).

    :param file: file to write to.
    :param task: name of the task
    :param time: time of the task in seconds
    :param task_label: label of the column containing task names
    :param time_label: label of the column containing time names
    :raise FileNotFoundError: when file with specified file name does
            not exist
    :raise AttributeError: when any of specified labels does not exist
            in the file

    """
    global arial_narrow_bold_font
    global arial_narrow_font

    if file is None or file == "":
        return

    # try opening Excel file.
    try:
        wb = openpyxl.load_workbook(file)
    except FileNotFoundError:
        raise

    ws = wb.active

    # find boundaries of non-empty cells in sheet `ws`
    max_r = ws.max_row
    max_c = ws.max_column

    # find in which columns task names and task times are stored
    try:
        task_column_letter = find_cell(ws, max_c, max_r, task_label).column_letter
        time_column_letter = find_cell(ws, max_c, max_r, time_label).column_letter
    except AttributeError:
        wb.close()
        raise

    # check if task with given name already exists
    existing_cell = find_cell(ws, max_c, max_r, task)

    if existing_cell is None:
        # if it doesn't exist, create it in new row
        new_task_cell = f"{task_column_letter}{max_r + 1}"
        new_time_cell = f"{time_column_letter}{max_r + 1}"
        ws[new_task_cell] = task
        ws[new_task_cell].font = arial_narrow_font
        ws[new_time_cell] = time / 3600  # write time in hours
        ws[new_time_cell].font = arial_narrow_font
        ws[new_time_cell].number_format = '0.00'
    else:
        # if it exists, add new time to existing time
        task_row = existing_cell.row
        time_cell = f"{time_column_letter}{task_row}"
        existing_value = ws[time_cell].value  # existing_value in hours
        time += existing_value * 3600  # time in seconds
        ws[time_cell] = time / 3600  # write time back in hours
        ws[time_cell].font = arial_narrow_font
        ws[time_cell].number_format = '0.00'

    wb.save(file)
    wb.close()


def list_existing_tasks(file: str, task_label: str) -> list[str]:
    """Returns the list of the names of the existing tasks in Excel file.

    :param file: file name
    :param task_label: label of the column with task names
    :return: list of the names of the existing tasks.
            Returns empty list if file name is not specified (empty).
    :raise FileNotFoundError: when file with specified file name does
            not exist
    :raise AttributeError: when specified label does not exist inside
            the file
    """

    if file is None or file == "":
        return []

    # try opening Excel file.
    try:
        wb = openpyxl.load_workbook(file)
    except FileNotFoundError:
        raise

    ws = wb.active
    all_tasks = []

    # find boundaries of non-empty cells in sheet `ws`
    max_r = ws.max_row
    max_c = ws.max_column

    # find cell with value `task_label`
    task_cell = find_cell(ws, max_c, max_r, task_label)

    if task_cell is None:
        wb.close()
        raise AttributeError

    start_row = task_cell.row + 1
    search_column = task_cell.column

    for row in range(start_row, max_r + 1):
        next_task = ws.cell(row=row, column=search_column).value
        all_tasks.append(next_task)

    wb.close()

    return all_tasks


if __name__ == '__main__':

    task_names = []
    start_time = None
    section_opened = False
    task_column_disabled = True
    time_column_disabled = True
    timer_stopped = True

    arial_narrow_bold_font = Font(name="Arial Narrow", size=10.5, bold=True)
    arial_narrow_font = Font(name="Arial Narrow", size=10.5)

    window = sg.Window('Time Your Work', layout, no_titlebar=True,
                       grab_anywhere=True, font='Arial 10', size=(520, 320),
                       finalize=True, use_default_focus=False,
                       margins=(15, 15))

    style = sg.ttk.Style()
    style.configure("TCombobox", borderwidth=0, relief='flat')
    style.configure('Vertical.TScrollbar', background="#ffffff", relief='flat',
                    borderwidth=0, troughcolor='#f2f2f2')

    while True:  # Event Loop
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':
            break

        if event == 'start':
            if timer_stopped:
                start_time = datetime.datetime.utcnow()
                local_timezone = get_localzone()
                timer_start = pytz.utc.localize(start_time).astimezone(local_timezone)
                timer_stopped = False
                window['-TIMER MESSAGE-'].update('timer started at')
                window['-TOTAL TIME HOURS-'].update(f'{timer_start.hour:d}:{timer_start.minute:02d}')

        if event == 'stop':
            if not timer_stopped:
                end_time = datetime.datetime.utcnow()
                total_time = round((end_time - start_time).total_seconds())
                timer_stopped = True
                window['-TOTAL TIME-'].update(total_time)
                window['-TIMER MESSAGE-'].update('total time')
                window['-TOTAL TIME HOURS-'].update(f'{total_time / 3600:.2f} h')

        if event == '-FILENAME-':
            file_name = str(values["-FILENAME-"])
            task_column = str(values['-TASK COLUMN-'])
            try:
                task_names = list_existing_tasks(file_name, task_column)
                window['-TASK NAMES-'].update(values=task_names)
            except FileNotFoundError:
                error_popup(f'ERROR. File {file_name} not found')
            except AttributeError:
                error_popup(f'ERROR. No "{task_column}" label in file {file_name}')

        if event.startswith('-OPEN SEC'):
            section_opened = not section_opened
            window['-OPEN SEC-'].update(SYMBOL_UP if section_opened else SYMBOL_DOWN)
            window['-SEC-'].update(visible=section_opened)

        if event == '-SET TASK COLUMN-':
            if not task_column_disabled:
                file_name = str(values['-FILENAME-'])
                task_column = str(values['-TASK COLUMN-'])
                try:
                    task_names = list_existing_tasks(file_name, task_column)
                    window['-TASK NAMES-'].update(values=task_names)
                except FileNotFoundError:
                    error_popup(f'ERROR. File {file_name} not found')
                    window['-TASK NAMES-'].update(values=[])
                except AttributeError:
                    error_popup(f'ERROR. No "{task_column}" label in file {file_name}')
                    window['-TASK NAMES-'].update(values=[])
            task_column_disabled = not task_column_disabled
            window['-SET TASK COLUMN-'].update(text='edit' if task_column_disabled else 'set')
            window['-TASK COLUMN-'].update(disabled=task_column_disabled)

        if event == '-SET TIME COLUMN-':
            time_column_disabled = not time_column_disabled
            window['-SET TIME COLUMN-'].update(text='edit' if time_column_disabled else 'set')
            window['-TIME COLUMN-'].update(disabled=time_column_disabled)

        if event == 'Save':
            file_name = str(values['-FILENAME-'])
            new_task = str(values['-TASK NAMES-'])
            new_time = int(window['-TOTAL TIME-'].get())
            task_column = str(values['-TASK COLUMN-'])
            time_column = str(values['-TIME COLUMN-'])
            try:
                write_data(file_name, new_task, new_time, task_column, time_column)
            except FileNotFoundError:
                error_popup(f'ERROR. File {file_name} not found')
            except AttributeError:
                error_popup(f'ERROR. No "{task_column}" and/or "{time_column}" label in file {file_name}')

        if event == 'New task':
            # reset timer
            start_time = None
            timer_stopped = True
            window['-TOTAL TIME-'].update(0)
            window['-TIMER MESSAGE-'].update('total time')
            window['-TOTAL TIME HOURS-'].update('0 h')
            # update existing task names list in Combobox -TASK NAMES-
            file_name = str(values["-FILENAME-"])
            task_column = str(values['-TASK COLUMN-'])
            try:
                task_names = list_existing_tasks(file_name, task_column)
                window['-TASK NAMES-'].update(values=task_names)
            except FileNotFoundError:
                error_popup(f'ERROR. File {file_name} not found')
            except AttributeError:
                error_popup(f'ERROR. No "{task_column}" label in file {file_name}')

    window.close()
